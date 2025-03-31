from openpyxl import load_workbook

class ExcelUtils:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active

    def setWs(self, ws):
        self.ws = self.wb[ws]

    def removeRows(self, rows):
        sorted(rows).reverse()
        for row in rows:
            self.ws.delete_rows(row)

    def fillCell(self, column_letter, row, value, wsIndex=None):
        ws = self.wb[wsIndex] if wsIndex != None else self.ws

        # Fill the data into the column
        ws[f"{column_letter}{row}"] = value

    def fillColumn(self, column_letter, data, start_row=2, wsIndex=None, wsNewName=None):
        ws = self.wb[wsIndex] if wsIndex != None else self.ws

        # Fill the data into the column
        for i, value in enumerate(data, start=start_row):
            ws[f"{column_letter}{i}"] = value
        if wsNewName:
            ws.title = wsNewName

    def save(self, file_path=None):
        if not file_path:
            file_path = self.file_path
        self.wb.save(file_path)
        self.wb.close()